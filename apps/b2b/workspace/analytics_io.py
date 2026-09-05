"""The report as a file, and as a line of text.

Export is what the download button on «Hisobotlar» hands back — the eight
figures, the per-employee table and the rows behind the first figure, as
XLSX or CSV, in the language the request came in. The same labels and number
formatting serve the subscription pass, which writes the figures into a chat
message and a mail body.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.b2b.workspace.analytics import Window

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CONTENT_TYPE = "text/csv; charset=utf-8"

# ─── Words ───────────────────────────────────────────────────────────────────
#
# (uz, ru) pairs. The phone has the same words in its own string tables; these
# are for the file and the message, which are written on the server.

SECTION_LABELS: dict[str, tuple[str, str]] = {
    "sales": ("Sotuv", "Продажи"),
    "tasks": ("Vazifalar", "Задачи"),
    "stock": ("Sklad", "Склад"),
    "trips": ("Komandirovka", "Командировки"),
    "attendance": ("Davomat", "Посещаемость"),
}

METRIC_LABELS: dict[str, dict[str, tuple[str, str]]] = {
    "sales": {
        "revenue": ("Daromad", "Выручка"),
        "avg_check": ("O'rtacha chek", "Средний чек"),
        "deals": ("Yopilgan bitimlar", "Закрытые сделки"),
        "conversion": ("Konversiya", "Конверсия"),
        "cycle": ("Bitim sikli", "Цикл сделки"),
        "response": ("Birinchi javob vaqti", "Время первого ответа"),
        "goods": ("Sotilgan tovarlar", "Проданные товары"),
        "profit": ("Valyuta profit", "Валютный профит"),
    },
    "tasks": {
        "completed": ("Bajarilgan", "Выполнено"),
        "created": ("Yaratilgan", "Создано"),
        "on_time": ("Muddatida", "В срок"),
        "late": ("Kechikib bajarilgan", "С опозданием"),
        "cycle": ("O'rtacha bajarish vaqti", "Среднее время выполнения"),
        "completion": ("Bajarilish darajasi", "Доля выполнения"),
        "open": ("Ochiq vazifalar", "Открытые задачи"),
        "overdue": ("Muddati o'tgan", "Просроченные"),
    },
    "stock": {
        "sold_value": ("Sotuv summasi", "Сумма продаж"),
        "sold_qty": ("Sotilgan miqdor", "Продано, шт"),
        "receipt_value": ("Kirim summasi", "Сумма поступлений"),
        "write_off_value": ("Hisobdan chiqarish", "Списания"),
        "movements": ("Harakatlar", "Движения"),
        "return_qty": ("Qaytarishlar", "Возвраты"),
        "stock_value": ("Qoldiq qiymati", "Стоимость остатков"),
        "low_stock": ("Kam qoldiq", "Низкий остаток"),
    },
    "trips": {
        "trips": ("Safarlar", "Поездки"),
        "travellers": ("Xodimlar", "Сотрудники"),
        "budget": ("Byudjet", "Бюджет"),
        "avg_budget": ("O'rtacha byudjet", "Средний бюджет"),
        "avg_days": ("O'rtacha davomiylik", "Средняя длительность"),
        "completed": ("Yakunlangan", "Завершено"),
        "cancelled": ("Bekor qilingan", "Отменено"),
        "active": ("Hozir safarda", "Сейчас в поездке"),
    },
    "attendance": {
        "rate": ("Davomat darajasi", "Посещаемость"),
        "present": ("Kelgan", "Присутствовали"),
        "absent": ("Kelmagan", "Отсутствовали"),
        "late": ("Kechikkan", "Опоздания"),
        "remote": ("Masofadan", "Удалённо"),
        "avg_checkin": ("O'rtacha kelish vaqti", "Среднее время прихода"),
        "avg_hours": ("O'rtacha ish soati", "Средние часы работы"),
        "people": ("Xodimlar", "Сотрудники"),
    },
}

COLUMN_LABELS: dict[str, tuple[str, str]] = {
    "revenue": ("Daromad", "Выручка"),
    "deals": ("Sotuv", "Сделки"),
    "avg_check": ("O'rtacha", "Средний"),
    "conversion": ("Konv.", "Конв."),
    "completed": ("Bajarildi", "Выполнено"),
    "on_time": ("Muddatida", "В срок"),
    "late": ("Kechikkan", "Опоздано"),
    "movements": ("Harakatlar", "Движения"),
    "sold_value": ("Sotuv", "Продажи"),
    "receipt_value": ("Kirim", "Приход"),
    "trips": ("Safarlar", "Поездки"),
    "days": ("Kunlar", "Дни"),
    "budget": ("Byudjet", "Бюджет"),
    "present": ("Kelgan", "Был"),
    "absent": ("Kelmagan", "Не был"),
}

_WORDS: dict[str, tuple[str, str]] = {
    "metric": ("Ko'rsatkich", "Показатель"),
    "value": ("Qiymat", "Значение"),
    "previous": ("Oldingi davr", "Прошлый период"),
    "change": ("O'zgarish", "Изменение"),
    "employee": ("Xodim", "Сотрудник"),
    "total": ("Jami", "Итого"),
    "records": ("Yozuvlar", "Записи"),
    "metrics": ("Ko'rsatkichlar", "Показатели"),
    "employees": ("Xodimlar bo'yicha", "По сотрудникам"),
    "title": ("Nomi", "Название"),
    "subtitle": ("Izoh", "Примечание"),
    "status": ("Holat", "Статус"),
    "amount": ("Summa", "Сумма"),
    "date": ("Sana", "Дата"),
    "period": ("Davr", "Период"),
    "report": ("Hisobot", "Отчёт"),
    "currency": ("so'm", "сум"),
    "million": ("mln", "млн"),
    "thousand": ("ming", "тыс"),
    "days": ("kun", "дн."),
    "hours": ("soat", "ч"),
    "pieces": ("ta", "шт"),
    "day": ("Kun", "День"),
    "week": ("Hafta", "Неделя"),
    "month": ("Oy", "Месяц"),
    "year": ("Yil", "Год"),
    "yesterday": ("kecha", "вчера"),
    "last_week": ("o'tgan hafta", "прошлая неделя"),
    "last_year": ("o'tgan yil", "прошлый год"),
}

_MONTHS_UZ = [
    "yanvar", "fevral", "mart", "aprel", "may", "iyun",
    "iyul", "avgust", "sentyabr", "oktyabr", "noyabr", "dekabr",
]
_MONTHS_RU = [
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]
_MONTHS_RU_NOM = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]


def _lang_index(lang: str) -> int:
    return 1 if (lang or "uz").lower().startswith("ru") else 0


def word(key: str, lang: str) -> str:
    return _WORDS.get(key, (key, key))[_lang_index(lang)]


def section_label(section: str, lang: str) -> str:
    return SECTION_LABELS.get(section, (section, section))[_lang_index(lang)]


def metric_label(section: str, key: str, lang: str) -> str:
    return METRIC_LABELS.get(section, {}).get(key, (key, key))[_lang_index(lang)]


def column_label(key: str, lang: str) -> str:
    return COLUMN_LABELS.get(key, (key, key))[_lang_index(lang)]


def month_name(month: int, lang: str, *, genitive: bool = True) -> str:
    if _lang_index(lang):
        return (_MONTHS_RU if genitive else _MONTHS_RU_NOM)[month - 1]
    return _MONTHS_UZ[month - 1]


# ─── Numbers ─────────────────────────────────────────────────────────────────

def _decimal_sep(value: str, lang: str) -> str:
    # Both languages write the decimal comma.
    return value.replace(".", ",")


def _group(value: float) -> str:
    whole = int(round(abs(value)))
    text = f"{whole:,}".replace(",", " ")
    return f"-{text}" if value < 0 else text


def _one_decimal(value: float, lang: str) -> str:
    text = f"{value:.0f}" if abs(value) >= 100 else f"{value:.1f}"
    return _decimal_sep(text, lang)


def compact_money(value: float, lang: str) -> str:
    """48 500 000 → «48,5 mln so'm» — what a card prints."""
    if abs(value) >= 1_000_000:
        return f"{_one_decimal(value / 1_000_000, lang)} {word('million', lang)} {word('currency', lang)}"
    if abs(value) >= 1_000:
        return f"{_one_decimal(value / 1_000, lang)} {word('thousand', lang)} {word('currency', lang)}"
    return f"{_group(value)} {word('currency', lang)}"


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_value(fmt: str, value: Any, lang: str) -> str:
    """One figure as a person reads it, in the file and in the message."""
    number = _to_float(value)
    if number is None:
        return "—"
    if fmt == "money":
        return compact_money(number, lang)
    if fmt == "percent":
        return f"{_one_decimal(number * 100, lang)}%"
    if fmt == "days":
        return f"{_one_decimal(number, lang)} {word('days', lang)}"
    if fmt == "hours":
        return f"{_one_decimal(number, lang)} {word('hours', lang)}"
    if fmt == "clock":
        minutes = int(round(number))
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    if fmt == "qty":
        return f"{_group(number)} {word('pieces', lang)}"
    return _group(number)


def format_change(change: Any, lang: str) -> str:
    """"+12%", "−3%" — whole percents, as the card prints them; a movement
    under one percent keeps a decimal so it does not read as none."""
    number = _to_float(change)
    if number is None:
        return "—"
    percent = number * 100
    sign = "+" if percent > 0 else ("−" if percent < 0 else "")
    if 0 < abs(percent) < 1:
        return f"{sign}{_decimal_sep(f'{abs(percent):.1f}', lang)}%"
    return f"{sign}{round(abs(percent))}%"


def _cell(fmt: str, value: Any) -> Any:
    """The raw figure for a spreadsheet cell — a number where there is one,
    so the reader can sum the column, and a clock string where there is not."""
    number = _to_float(value)
    if number is None:
        return None
    if fmt == "percent":
        return round(number * 100, 1)
    if fmt == "clock":
        minutes = int(round(number))
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    if fmt == "money":
        return round(number, 2)
    if fmt == "count":
        return int(round(number))
    return round(number, 2)


# ─── Period words ────────────────────────────────────────────────────────────

def window_label(window: Window, lang: str) -> str:
    """«1–30 avgust 2026», «12 avgust 2026», «Avgust 2026», «2026» — the
    subtitle under the controls and the first line of the message."""
    start = window.start.astimezone(window.start.tzinfo)
    start_d = window.start_date
    last = window.end_date.toordinal() - 1
    end_d = date.fromordinal(last) if last >= start_d.toordinal() else start_d
    ru = _lang_index(lang) == 1
    if window.period == "day":
        return f"{start_d.day} {month_name(start_d.month, lang)} {start_d.year}"
    if window.period == "year":
        return f"{start_d.year}"
    if start_d.month == end_d.month:
        return f"{start_d.day}–{end_d.day} {month_name(start_d.month, lang)} {start_d.year}"
    return (
        f"{start_d.day} {month_name(start_d.month, lang)} – "
        f"{end_d.day} {month_name(end_d.month, lang)} {end_d.year}"
    )


def compare_label(window: Window, lang: str) -> str:
    """What the comparison figure is: «iyul», «kecha», «o'tgan hafta»."""
    if window.period == "day":
        return word("yesterday", lang)
    if window.period == "week":
        return word("last_week", lang)
    if window.period == "month":
        d = window.compare_start_date
        return month_name(d.month, lang, genitive=False)
    return str(window.compare_start_date.year)


def period_word(period: str, lang: str) -> str:
    return word(period, lang)


# ─── Files ───────────────────────────────────────────────────────────────────

def _rows_metrics(section: str, report: dict[str, Any], window: Window, lang: str) -> tuple[list[str], list[list[Any]]]:
    header = [word("metric", lang), word("value", lang), word("previous", lang), word("change", lang)]
    rows = []
    for metric in report["metrics"]:
        rows.append([
            metric_label(section, metric["key"], lang),
            _cell(metric["format"], metric["value"]),
            _cell(metric["format"], metric.get("previous")),
            _cell("percent", metric.get("change")),
        ])
    return header, rows


def _rows_employees(report: dict[str, Any], lang: str) -> tuple[list[str], list[list[Any]]]:
    table = report["employees"]
    columns = table["columns"]
    header = [word("employee", lang)] + [column_label(c["key"], lang) for c in columns] + [word("change", lang)]
    rows = []
    for row in table["rows"]:
        rows.append(
            [row["full_name"]]
            + [_cell(c["format"], row["values"].get(c["key"])) for c in columns]
            + [_cell("percent", row.get("change"))]
        )
    total = table.get("total") or {}
    if total:
        rows.append(
            [word("total", lang)]
            + [_cell(c["format"], total.get(c["key"])) for c in columns]
            + [_cell("percent", total.get("change"))]
        )
    return header, rows


def _rows_items(items: dict[str, Any], lang: str) -> tuple[list[str], list[list[Any]]]:
    header = [
        word("title", lang), word("subtitle", lang), word("status", lang),
        word("amount", lang), word("date", lang), word("employee", lang),
    ]
    rows = []
    for row in items.get("rows", []):
        stamp = row.get("date")
        rows.append([
            row.get("title"), row.get("subtitle"), row.get("status"),
            _cell("money", row.get("amount")) if row.get("amount") is not None else None,
            stamp, row.get("employee"),
        ])
    return header, rows


def _sheet(book: Workbook, title: str, header: Sequence[str], rows: Sequence[Sequence[Any]], *, first: bool = False):
    sheet = book.active if first else book.create_sheet()
    sheet.title = title[:31]
    sheet.append(list(header))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([
            float(v) if isinstance(v, Decimal) else
            v.replace(tzinfo=None) if isinstance(v, datetime) else v
            for v in row
        ])
    for index, _ in enumerate(header, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = 22
    return sheet


def file_name(section: str, window: Window, fmt: str) -> str:
    return f"weel-{section}-{window.period}-{window.start_date:%Y%m%d}.{fmt}"


def export_xlsx(section: str, report: dict[str, Any], items: dict[str, Any], window: Window, lang: str) -> bytes:
    book = Workbook()
    header, rows = _rows_metrics(section, report, window, lang)
    sheet = _sheet(book, word("metrics", lang), header, rows, first=True)
    # The period the file describes, above the table it describes.
    sheet.insert_rows(1, amount=2)
    sheet["A1"] = f"{word('report', lang)}: {section_label(section, lang)}"
    sheet["A1"].font = Font(bold=True)
    sheet["A2"] = f"{word('period', lang)}: {period_word(window.period, lang)} · {window_label(window, lang)}"
    for cell in sheet[3]:
        cell.font = Font(bold=True)
    header, rows = _rows_employees(report, lang)
    _sheet(book, word("employees", lang), header, rows)
    header, rows = _rows_items(items, lang)
    _sheet(book, word("records", lang), header, rows)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def export_csv(section: str, report: dict[str, Any], items: dict[str, Any], window: Window, lang: str) -> bytes:
    """One file, three blocks. UTF-8 with a BOM and semicolons, which is what
    Excel on a Windows laptop in Tashkent opens without a wizard."""
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";", lineterminator="\r\n")
    writer.writerow([f"{word('report', lang)}: {section_label(section, lang)}"])
    writer.writerow([f"{word('period', lang)}: {period_word(window.period, lang)} · {window_label(window, lang)}"])
    writer.writerow([])
    for header, rows in (
        _rows_metrics(section, report, window, lang),
        _rows_employees(report, lang),
        _rows_items(items, lang),
    ):
        writer.writerow(header)
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])
        writer.writerow([])
    return ("﻿" + out.getvalue()).encode("utf-8")


# ─── Text ────────────────────────────────────────────────────────────────────

def summary_text(section: str, report: dict[str, Any], window: Window, lang: str) -> str:
    """The figures as a chat message: one line per metric, the change beside
    it, the period on top."""
    lines = [
        f"📊 {section_label(section, lang)} — {period_word(window.period, lang).lower()} · {window_label(window, lang)}",
    ]
    compare = compare_label(window, lang)
    for metric in report["metrics"]:
        label = metric_label(section, metric["key"], lang)
        value = format_value(metric["format"], metric["value"], lang)
        if metric.get("snapshot"):
            lines.append(f"• {label}: {value}")
            continue
        change = metric.get("change")
        tail = f" ({format_change(change, lang)}, {compare})" if change is not None else ""
        lines.append(f"• {label}: {value}{tail}")
    return "\n".join(lines)
