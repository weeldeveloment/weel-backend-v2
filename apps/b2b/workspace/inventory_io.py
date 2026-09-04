"""XLSX in and out of the stock room.

Export is the catalogue, the balances or the ledger as a spreadsheet.
Import is a spreadsheet of products read twice: once to *preview* — every
row labelled create / update / error, nothing written — and once to
*commit* what the person confirmed, matching on SKU so the same file
uploaded twice updates rather than duplicates.
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.b2b.workspace import inventory_repository as inventory
from apps.b2b.workspace.inventory_repository import InventoryError, _q


# ─── Export ───────────────────────────────────────────────────────────────────

def _sheet(title: str, header: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = title[:31]
    sheet.append(list(header))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([
            float(v) if isinstance(v, Decimal) else
            v.replace(tzinfo=None) if isinstance(v, datetime) else v
            for v in row
        ])
    for index, _ in enumerate(header, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = 18
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def export_catalog(company_id: int, *, with_costs: bool) -> bytes:
    products = inventory.list_products(company_id, include_variants=True, limit=100000)
    header = [
        "Nomi", "Turi", "Artikul", "Shtrix-kod", "Kategoriya", "Brend", "Yetkazib beruvchi",
        "O'lchov", "Sotuv narxi", "Ulgurji narx", "Minimal qoldiq", "Qoldiq", "Rezerv", "Mavjud",
        "Holat", "Variant", "Izoh",
    ]
    if with_costs:
        header.insert(8, "Xarid narxi")
    rows = []
    for p in products:
        row = [
            p["name"], p["kind"], p.get("sku"), p.get("barcode"), p.get("category_name"),
            p.get("brand"), p.get("supplier_name"), p.get("unit"), _q(p.get("sale_price")),
            _q(p.get("wholesale_price")), _q(p.get("min_stock")), p["stock_total"],
            p["reserved"], p["available"], p["stock_status"], p.get("variant_label"),
            p.get("description"),
        ]
        if with_costs:
            row.insert(8, _q(p.get("purchase_price")))
        rows.append(row)
    return _sheet("Katalog", header, rows)


def export_stock(company_id: int, *, with_costs: bool) -> bytes:
    products = inventory.list_products(company_id, include_variants=True, limit=100000)
    header = ["Tovar", "Artikul", "Sklad", "Qoldiq", "O'lchov"]
    if with_costs:
        header += ["Xarid narxi", "Qiymati"]
    rows = []
    for p in products:
        for s in p["stocks"]:
            row = [p["name"], p.get("sku"), s["warehouse_name"], s["quantity"], p.get("unit")]
            if with_costs:
                row += [_q(p.get("purchase_price")), s["quantity"] * _q(p.get("purchase_price"))]
            rows.append(row)
    return _sheet("Qoldiqlar", header, rows)


def export_movements(company_id: int, *, with_costs: bool, **filters: Any) -> bytes:
    movements = inventory.list_movements(company_id, limit=100000, **filters)
    header = [
        "ID", "Sana", "Turi", "Hujjat", "Tovar", "Artikul", "Sklad", "Qayerga", "Miqdor", "O'lchov",
        "Narx", "Summa", "Valyuta", "Kim", "Yetkazib beruvchi", "Mijoz", "Izoh",
    ]
    if with_costs:
        header.insert(12, "Tannarx")
    rows = []
    for m in movements:
        qty = _q(m["quantity"])
        row = [
            m["id"], m["created_at"], m["kind"], m.get("document_number"), m.get("product_name"),
            m.get("sku"), m.get("warehouse_name"), m.get("to_warehouse_name"), qty, m.get("unit"),
            _q(m.get("unit_cost")), abs(qty) * _q(m.get("unit_cost")), m.get("currency"),
            m.get("author_name"), m.get("supplier_name"), m.get("customer_name"), m.get("note"),
        ]
        if with_costs:
            row.insert(12, _q(m.get("cost_price")))
        rows.append(row)
    return _sheet("Harakatlar", header, rows)


# ─── Import ───────────────────────────────────────────────────────────────────

#: Every spelling a column header may come in, mapped to the field it fills.
_HEADERS = {
    "name": ("nomi", "nom", "name", "tovar", "наименование", "название", "товар"),
    "sku": ("artikul", "sku", "артикул"),
    "barcode": ("shtrix-kod", "shtrix kod", "shtrixkod", "barcode", "штрих-код", "штрихкод"),
    "category": ("kategoriya", "category", "категория"),
    "brand": ("brend", "brand", "бренд"),
    "supplier": ("yetkazib beruvchi", "supplier", "поставщик"),
    "purchase_price": ("xarid narxi", "tannarx", "purchase_price", "purchase price", "закупочная цена", "себестоимость"),
    "sale_price": ("sotuv narxi", "sale_price", "sale price", "narx", "цена продажи", "цена"),
    "wholesale_price": ("ulgurji narx", "wholesale_price", "оптовая цена"),
    "unit": ("o'lchov", "olchov", "birlik", "unit", "ед.", "единица"),
    "min_stock": ("minimal qoldiq", "min_stock", "мин. остаток", "минимальный остаток"),
    "initial_quantity": ("qoldiq", "boshlang'ich qoldiq", "boshlangich qoldiq", "miqdor", "quantity", "initial_quantity", "остаток", "количество"),
    "warehouse": ("sklad", "warehouse", "склад"),
    "kind": ("turi", "kind", "тип"),
    "description": ("izoh", "tavsif", "description", "описание"),
}


def _column_map(header_row: Sequence[Any]) -> dict[str, int]:
    found: dict[str, int] = {}
    for index, raw in enumerate(header_row):
        label = str(raw or "").strip().lower()
        if not label:
            continue
        for field, spellings in _HEADERS.items():
            if field not in found and label in spellings:
                found[field] = index
                break
    return found


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(str(value))


def parse_import(company_id: int, file_bytes: bytes) -> dict[str, Any]:
    """Reads the spreadsheet and says what importing it would do.

    Every row comes back as it will be written, with ``action`` create /
    update / error and the messages behind an error. Nothing is stored.
    """
    try:
        book = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InventoryError(f"Faylni o'qib bo'lmadi: {exc}", code="bad_file")
    sheet = book.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        raise InventoryError("Fayl bo'sh.", code="bad_file")
    columns = _column_map(header)
    if "name" not in columns:
        raise InventoryError(
            "\"Nomi\" ustuni topilmadi. Birinchi qatorda ustun nomlari bo'lishi kerak.", code="bad_file"
        )
    rows: list[dict[str, Any]] = []
    seen_skus: set[str] = set()
    for line_no, values in enumerate(iterator, start=2):
        values = list(values or ())
        get = lambda field: (values[columns[field]] if field in columns and columns[field] < len(values) else None)  # noqa: E731
        name = str(get("name") or "").strip()
        if not name and not any(str(v or "").strip() for v in values):
            continue
        errors: list[str] = []
        if not name:
            errors.append("Nomi bo'sh")
        sku = str(get("sku") or "").strip() or None
        barcode = str(get("barcode") or "").strip() or None
        if barcode and barcode.endswith(".0"):
            barcode = barcode[:-2]
        numbers: dict[str, Decimal | None] = {}
        for field in ("purchase_price", "sale_price", "wholesale_price", "min_stock", "initial_quantity"):
            try:
                numbers[field] = _decimal(get(field))
            except ValueError as exc:
                errors.append(f"{field}: '{exc}' raqam emas")
                numbers[field] = None
        kind = str(get("kind") or "product").strip().lower()
        kind = {"tovar": "product", "xizmat": "service", "komplekt": "bundle",
                "товар": "product", "услуга": "service", "комплект": "bundle"}.get(kind, kind)
        if kind not in ("product", "service", "bundle"):
            kind = "product"
        existing = inventory.find_product_by_sku(company_id, sku) if sku else None
        if sku and sku in seen_skus:
            errors.append("Artikul faylda takrorlangan")
        if sku:
            seen_skus.add(sku)
        if barcode:
            taken = inventory._sku_taken(company_id, "barcode", barcode, exclude_id=existing["id"] if existing else None)
            if taken:
                errors.append("Bu shtrix-kod boshqa tovarda")
        rows.append({
            "line": line_no,
            "action": "error" if errors else ("update" if existing else "create"),
            "errors": errors,
            "existing_id": existing["id"] if existing else None,
            "name": name,
            "kind": kind,
            "sku": sku,
            "barcode": barcode,
            "category": str(get("category") or "").strip() or None,
            "brand": str(get("brand") or "").strip() or None,
            "supplier": str(get("supplier") or "").strip() or None,
            "unit": str(get("unit") or "").strip() or None,
            "description": str(get("description") or "").strip() or None,
            "warehouse": str(get("warehouse") or "").strip() or None,
            **numbers,
        })
    return {
        "columns": sorted(columns),
        "rows": rows,
        "create_count": sum(1 for r in rows if r["action"] == "create"),
        "update_count": sum(1 for r in rows if r["action"] == "update"),
        "error_count": sum(1 for r in rows if r["action"] == "error"),
    }


def commit_import(
    company_id: int,
    rows: Sequence[dict[str, Any]],
    *,
    author_id: int | None,
    update_existing: bool = True,
    update_fields: Sequence[str] | None = None,
) -> dict[str, Any]:
    """Writes the rows a preview produced and the person confirmed.

    Matching is by SKU. An existing product gets only the fields named in
    ``update_fields`` (all of them when None) and never a second opening
    balance; a new one gets everything, and its balance as a receipt.
    """
    created, updated, skipped, errors = 0, 0, 0, []
    fields = set(update_fields or (
        "name", "barcode", "category", "brand", "supplier", "unit",
        "purchase_price", "sale_price", "wholesale_price", "min_stock", "description",
    ))
    warehouses = {w["name"].strip().lower(): w for w in inventory.list_warehouses(company_id)}
    for row in rows:
        if row.get("action") == "error" or not (row.get("name") or "").strip():
            skipped += 1
            continue
        try:
            category = inventory.find_or_create_category(company_id, row.get("category") or "")
            supplier = inventory.find_or_create_supplier(company_id, row.get("supplier") or "")
            sku = (row.get("sku") or "").strip() or None
            existing = inventory.find_product_by_sku(company_id, sku) if sku else None
            if existing:
                if not update_existing:
                    skipped += 1
                    continue
                changes: dict[str, Any] = {}
                if "name" in fields:
                    changes["name"] = row["name"].strip()
                if "barcode" in fields and row.get("barcode"):
                    changes["barcode"] = row["barcode"]
                if "category" in fields and category:
                    changes["category_id"] = category["id"]
                if "supplier" in fields and supplier:
                    changes["supplier_id"] = supplier["id"]
                if "brand" in fields and row.get("brand"):
                    changes["brand"] = row["brand"]
                if "unit" in fields and row.get("unit"):
                    changes["unit"] = row["unit"]
                if "description" in fields and row.get("description"):
                    changes["description"] = row["description"]
                for price in ("purchase_price", "sale_price", "wholesale_price", "min_stock"):
                    if price in fields and row.get(price) is not None:
                        changes[price] = row[price]
                inventory.update_product(existing["id"], company_id, author_id=author_id, **changes)
                updated += 1
            else:
                warehouse = warehouses.get((row.get("warehouse") or "").strip().lower())
                inventory.create_product(
                    company_id,
                    name=row["name"].strip(),
                    author_id=author_id,
                    kind=row.get("kind") or "product",
                    category_id=category["id"] if category else None,
                    supplier_id=supplier["id"] if supplier else None,
                    brand=row.get("brand"),
                    sku=sku,
                    barcode=row.get("barcode"),
                    unit=row.get("unit") or "dona",
                    purchase_price=row.get("purchase_price") or 0,
                    sale_price=row.get("sale_price"),
                    wholesale_price=row.get("wholesale_price") or 0,
                    min_stock=row.get("min_stock") or 0,
                    description=row.get("description"),
                    initial_quantity=row.get("initial_quantity"),
                    initial_warehouse_id=warehouse["id"] if warehouse else None,
                )
                created += 1
        except InventoryError as exc:
            errors.append({"line": row.get("line"), "name": row.get("name"), "error": str(exc)})
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}
