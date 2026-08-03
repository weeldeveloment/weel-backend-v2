from __future__ import annotations

import io
import logging
from pathlib import PurePosixPath
from decimal import Decimal
from typing import Any

from django.core.files.storage import default_storage
from django.db import IntegrityError
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework import status
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from apps.platform.authentication import PmsJWTAuthentication
from apps.platform.raw_repository import get_organization_by_id, get_organization_member
from apps.shared.permissions import HasOrganization

from apps.pms.repository import (
    _generate_voucher_number,
    accept_booking,
    add_property_image,
    block_dates,
    cancel_booking,
    change_meal_plan,
    check_in_booking,
    check_out_booking,
    complain_review,
    create_booking,
    create_guest,
    create_property,
    create_review,
    create_room,
    create_room_type,
    delete_property,
    delete_property_image,
    delete_room,
    delete_room_type,
    expire_holds,
    find_or_create_guest,
    get_analytics,
    get_booking,
    get_booking_history,
    get_guest,
    get_or_create_calendar_slot,
    get_property,
    get_property_images,
    get_room,
    get_room_availability,
    get_room_type,
    hold_dates,
    list_bookings,
    list_guests,
    list_properties,
    list_reviews,
    list_room_types,
    list_rooms,
    mass_update_rooms,
    move_booking,
    respond_to_review,
    unblock_dates,
    unhold_dates,
    update_booking,
    update_booking_with_guest,
    update_guest,
    update_property,
    update_room,
    update_room_type,
)
from apps.pms.serializers import (
    AnalyticsQuerySerializer,
    AnalyticsResponseSerializer,
    BookingHistorySerializer,
    BookingSerializer,
    CalendarSlotSerializer,
    DateRangeSerializer,
    GuestSerializer,
    MealPlanChangeSerializer,
    MoveBookingSerializer,
    PropertyImageSerializer,
    PropertySerializer,
    ResizeBookingSerializer,
    ReviewComplainSerializer,
    ReviewRespondSerializer,
    ReviewSerializer,
    RoomIdsSerializer,
    RoomMassUpdateItemSerializer,
    RoomSerializer,
    RoomTypeSerializer,
)

logger = logging.getLogger(__name__)


def _delete_room_storage(prefix: str) -> None:
    try:
        directories, files = default_storage.listdir(prefix)
    except FileNotFoundError:
        return
    except Exception:
        logger.exception("Failed to list room storage prefix %s", prefix)
        return

    for filename in files:
        path = str(PurePosixPath(prefix, filename))
        try:
            default_storage.delete(path)
        except Exception:
            logger.exception("Failed to delete room image %s", path)

    for directory in directories:
        _delete_room_storage(str(PurePosixPath(prefix, directory)))


class PMSBaseView(APIView):
    authentication_classes = [PmsJWTAuthentication]
    permission_classes = [IsAuthenticated, HasOrganization]


def _get_user_id(request) -> int | None:
    user = getattr(request, "user", None)
    if isinstance(user, dict):
        return user.get("id")
    return getattr(user, "id", None)


def _require_org(request):
    org = getattr(request, "organization", None)
    org_id = None
    if org:
        if isinstance(org, dict):
            org_id = org.get("id")
        else:
            org_id = org

    user = getattr(request, "user", None)
    if org_id is None:
        if isinstance(user, dict):
            org_id = user.get("organization_id")
        else:
            org_id = getattr(user, "organization_id", None)

    user_id = _get_user_id(request)
    try:
        parsed_org_id = int(org_id) if org_id is not None else None
    except (TypeError, ValueError):
        return None

    if not parsed_org_id or not user_id:
        return None

    if not get_organization_by_id(parsed_org_id):
        return None

    if not get_organization_member(parsed_org_id, int(user_id)):
        return None

    return parsed_org_id


class PropertyListCreateView(PMSBaseView):
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @swagger_auto_schema(responses={200: PropertySerializer(many=True)})
    def get(self, request):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        props = list_properties(organization_id=int(org_id))
        return Response(PropertySerializer(props, many=True).data)

    @swagger_auto_schema(
        responses={201: PropertySerializer()},
        operation_description="Create a new property",
        request_body=PropertySerializer,
    )
    def post(self, request):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = PropertySerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            prop = create_property(organization_id=int(org_id), **serializer.validated_data)
        except IntegrityError as e:
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if not prop:
            return Response({"detail": "Failed to create property."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(PropertySerializer(prop).data, status=status.HTTP_201_CREATED)


class PropertyRetrieveUpdateDestroyView(PMSBaseView):
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @swagger_auto_schema(responses={200: PropertySerializer()})
    def get(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(PropertySerializer(prop).data)

    @swagger_auto_schema(responses={200: PropertySerializer()})
    def patch(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = PropertySerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        prop = update_property(property_id, organization_id=int(org_id), **serializer.validated_data)
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(PropertySerializer(prop).data)

    @swagger_auto_schema(responses={204: "Deleted"})
    def delete(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        if not delete_property(property_id, organization_id=int(org_id)):
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class PropertyImageCreateView(PMSBaseView):
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(responses={201: PropertyImageSerializer()})
    def post(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        image_file = request.FILES.get("image")
        if not image_file:
            return Response({"detail": "No image file provided."}, status=status.HTTP_400_BAD_REQUEST)

        path = default_storage.save(f"pms/properties/{property_id}/{image_file.name}", image_file)
        image_url = default_storage.url(path)

        order = request.data.get("order", 0)
        img = add_property_image(property_id, image_url, int(order))
        return Response(PropertyImageSerializer(img).data, status=status.HTTP_201_CREATED)


class PropertyImageDeleteView(PMSBaseView):
    @swagger_auto_schema(responses={204: "Deleted"})
    def delete(self, request, property_id, image_id):
        if not delete_property_image(image_id):
            return Response({"detail": "Image not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class RoomImageCreateView(PMSBaseView):
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        responses={201: openapi.Response("Image uploaded", schema=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={"image_url": openapi.Schema(type=openapi.TYPE_STRING)},
        ))},
        manual_parameters=[openapi.Parameter("image", openapi.IN_FORM, type=openapi.TYPE_FILE, required=True)],
    )
    def post(self, request, property_id, room_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        room = get_room(room_id, property_id)
        if not room:
            return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)

        image_file = request.FILES.get("image")
        if not image_file:
            return Response({"detail": "No image file provided."}, status=status.HTTP_400_BAD_REQUEST)

        path = default_storage.save(f"pms/properties/{property_id}/rooms/{room_id}/{image_file.name}", image_file)
        image_url = default_storage.url(path)
        return Response({"image_url": image_url}, status=status.HTTP_201_CREATED)


class RoomListCreateView(PMSBaseView):
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @swagger_auto_schema(responses={200: RoomSerializer(many=True)})
    def get(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        room_type_name = request.query_params.get("room_type_name")
        rooms = list_rooms(property_id, room_type_name=room_type_name if room_type_name else None)
        return Response(RoomSerializer(rooms, many=True).data)

    @swagger_auto_schema(responses={201: RoomSerializer()})
    def post(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = RoomSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            room = create_room(property_id=property_id, **serializer.validated_data)
        except IntegrityError as e:
            if "duplicate key" in str(e) and "room_number" in str(e):
                return Response({"room_number": "A room with this number already exists for this property."}, status=status.HTTP_400_BAD_REQUEST)
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if not room:
            return Response({"detail": "Failed to create room."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(RoomSerializer(room).data, status=status.HTTP_201_CREATED)


class RoomRetrieveUpdateDestroyView(PMSBaseView):
    @swagger_auto_schema(responses={200: RoomSerializer()})
    def get(self, request, property_id, room_id):
        room = get_room(room_id, property_id)
        if not room:
            return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RoomSerializer(room).data)

    @swagger_auto_schema(responses={200: RoomSerializer()})
    def patch(self, request, property_id, room_id):
        serializer = RoomSerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        room = update_room(room_id, **serializer.validated_data)
        if not room:
            return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RoomSerializer(room).data)

    @swagger_auto_schema(
        responses={
            204: "Deleted",
            404: "Room not found",
            409: "Room has booking history",
        }
    )
    def delete(self, request, property_id, room_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)

        prop = get_property(property_id, organization_id=int(org_id))
        if not prop or not get_room(room_id, property_id):
            return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            result = delete_room(room_id, property_id)
        except IntegrityError:
            result = "has_bookings"

        if result == "has_bookings":
            return Response(
                {
                    "code": "room_has_bookings",
                    "detail": "Room cannot be deleted because it has booking history.",
                },
                status=status.HTTP_409_CONFLICT,
            )
        if result == "not_found":
            return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)

        _delete_room_storage(f"pms/properties/{property_id}/rooms/{room_id}")
        return Response(status=status.HTTP_204_NO_CONTENT)


class RoomMassUpdateView(PMSBaseView):
    @swagger_auto_schema(responses={200: RoomSerializer(many=True)})
    def post(self, request, property_id):
        if not isinstance(request.data, list):
            return Response({"detail": "Expected a list of room updates."}, status=status.HTTP_400_BAD_REQUEST)

        results = []
        for item in request.data:
            item_serializer = RoomMassUpdateItemSerializer(data=item)
            if not item_serializer.is_valid():
                return Response(item_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            room = update_room(item_serializer.validated_data["id"], **{
                k: v for k, v in item_serializer.validated_data.items() if k != "id"
            })
            if room:
                results.append(RoomSerializer(room).data)

        return Response(results, status=status.HTTP_200_OK)


class RoomTypeListCreateView(PMSBaseView):
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @swagger_auto_schema(responses={200: RoomTypeSerializer(many=True)})
    def get(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)
        room_types = list_room_types(property_id)
        return Response(RoomTypeSerializer(room_types, many=True).data)

    @swagger_auto_schema(responses={201: RoomTypeSerializer()})
    def post(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)
        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = RoomTypeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        room_type = create_room_type(property_id=property_id, **serializer.validated_data)
        if not room_type:
            return Response({"detail": "Failed to create room type."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(RoomTypeSerializer(room_type).data, status=status.HTTP_201_CREATED)


class RoomTypeRetrieveUpdateDestroyView(PMSBaseView):
    @swagger_auto_schema(responses={200: RoomTypeSerializer()})
    def get(self, request, property_id, room_type_id):
        room_type = get_room_type(room_type_id, property_id)
        if not room_type:
            return Response({"detail": "Room type not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RoomTypeSerializer(room_type).data)

    @swagger_auto_schema(responses={200: RoomTypeSerializer()})
    def patch(self, request, property_id, room_type_id):
        serializer = RoomTypeSerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        room_type = update_room_type(room_type_id, **serializer.validated_data)
        if not room_type:
            return Response({"detail": "Room type not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RoomTypeSerializer(room_type).data)

    @swagger_auto_schema(responses={204: "Deleted"})
    def delete(self, request, property_id, room_type_id):
        if not delete_room_type(room_type_id):
            return Response({"detail": "Room type not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class CalendarView(PMSBaseView):
    @swagger_auto_schema(responses={200: CalendarSlotSerializer(many=True)})
    def get(self, request, property_id):
        serializer = DateRangeSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        from_date = serializer.validated_data["from_date"]
        to_date = serializer.validated_data["to_date"]
        room_id = request.query_params.get("room_id")

        slots = get_room_availability(property_id=property_id, from_date=from_date, to_date=to_date)
        return Response(slots)


class CalendarBlockView(PMSBaseView):
    @swagger_auto_schema(responses={201: CalendarSlotSerializer(many=True)})
    def post(self, request, property_id):
        serializer = RoomIdsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        slots = block_dates(
            room_ids=serializer.validated_data["room_ids"],
            from_date=serializer.validated_data["from_date"],
            to_date=serializer.validated_data["to_date"],
        )
        return Response(CalendarSlotSerializer(slots, many=True).data, status=status.HTTP_201_CREATED)


class CalendarUnblockView(PMSBaseView):
    @swagger_auto_schema(responses={200: openapi.Response("Unblocked count", openapi.Schema(type=openapi.TYPE_OBJECT, properties={"unblocked": openapi.Schema(type=openapi.TYPE_INTEGER)}))})
    def post(self, request, property_id):
        serializer = RoomIdsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        count = unblock_dates(
            room_ids=serializer.validated_data["room_ids"],
            from_date=serializer.validated_data["from_date"],
            to_date=serializer.validated_data["to_date"],
        )
        return Response({"unblocked": count})


class CalendarHoldView(PMSBaseView):
    @swagger_auto_schema(responses={201: CalendarSlotSerializer(many=True)})
    def post(self, request, property_id):
        serializer = RoomIdsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        slots = hold_dates(
            room_ids=serializer.validated_data["room_ids"],
            from_date=serializer.validated_data["from_date"],
            to_date=serializer.validated_data["to_date"],
            hold_duration_minutes=serializer.validated_data.get("hold_duration_minutes", 30),
        )
        return Response(CalendarSlotSerializer(slots, many=True).data, status=status.HTTP_201_CREATED)


class CalendarUnholdView(PMSBaseView):
    @swagger_auto_schema(responses={200: openapi.Response("Unheld count", openapi.Schema(type=openapi.TYPE_OBJECT, properties={"unheld": openapi.Schema(type=openapi.TYPE_INTEGER)}))})
    def post(self, request, property_id):
        serializer = RoomIdsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        count = unhold_dates(
            room_ids=serializer.validated_data["room_ids"],
            from_date=serializer.validated_data["from_date"],
            to_date=serializer.validated_data["to_date"],
        )
        return Response({"unheld": count})


class GuestListCreateView(PMSBaseView):
    @swagger_auto_schema(responses={200: GuestSerializer(many=True)})
    def get(self, request):
        search = request.query_params.get("search")
        guests = list_guests(search=search)
        return Response(GuestSerializer(guests, many=True).data)

    @swagger_auto_schema(responses={201: GuestSerializer()})
    def post(self, request):
        serializer = GuestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            guest = create_guest(**serializer.validated_data)
        except IntegrityError as e:
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if not guest:
            return Response({"detail": "Failed to create guest."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(GuestSerializer(guest).data, status=status.HTTP_201_CREATED)


class GuestRetrieveUpdateView(PMSBaseView):
    @swagger_auto_schema(responses={200: GuestSerializer()})
    def get(self, request, guest_id):
        guest = get_guest(guest_id)
        if not guest:
            return Response({"detail": "Guest not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(GuestSerializer(guest).data)

    @swagger_auto_schema(responses={200: GuestSerializer()})
    def patch(self, request, guest_id):
        serializer = GuestSerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        guest = update_guest(guest_id, **serializer.validated_data)
        if not guest:
            return Response({"detail": "Guest not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(GuestSerializer(guest).data)


class BookingListCreateView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer(many=True)})
    def get(self, request, property_id):
        status_filter = request.query_params.get("status")
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")
        room_id = request.query_params.get("room_id")

        bookings = list_bookings(
            property_id=property_id,
            status=status_filter,
            from_date=from_date,
            to_date=to_date,
            room_id=int(room_id) if room_id else None,
        )
        return Response(BookingSerializer(bookings, many=True).data)

    @swagger_auto_schema(responses={201: BookingSerializer()})
    def post(self, request, property_id):
        serializer = BookingSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated = serializer.validated_data
        guest_id = validated.pop("guest_id", None)

        if not guest_id:
            guest_data = request.data.get("guest", {})
            if guest_data:
                guest = find_or_create_guest(
                    first_name=guest_data.get("first_name", "Guest"),
                    last_name=guest_data.get("last_name"),
                    email=guest_data.get("email"),
                    phone=guest_data.get("phone"),
                )
                guest_id = guest.get("id")

        try:
            booking = create_booking(
                property_id=property_id,
                check_in=validated["check_in"],
                check_out=validated["check_out"],
                room_id=validated["room_id"],
                guest_id=guest_id,
                created_by=_get_user_id(request),
                **{k: v for k, v in validated.items() if k not in ("check_in", "check_out", "room_id")},
            )
        except IntegrityError as e:
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if not booking:
            return Response({"detail": "Failed to create booking."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(BookingSerializer(booking).data, status=status.HTTP_201_CREATED)


class BookingRetrieveView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def get(self, request, property_id, booking_id):
        booking = get_booking(booking_id, property_id)
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(BookingSerializer(booking).data)

    @swagger_auto_schema(
        request_body=BookingSerializer,
        responses={200: BookingSerializer()},
    )
    def patch(self, request, property_id, booking_id):
        booking = get_booking(booking_id, property_id)
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = BookingSerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated = serializer.validated_data
        room_id = validated.get("room_id")
        if room_id is not None and not get_room(room_id, property_id):
            return Response(
                {"room_id": "Room not found for this property."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        updated = update_booking_with_guest(
            booking_id,
            guest_first_name=request.data.get("guest_first_name"),
            guest_last_name=request.data.get("guest_last_name"),
            user_id=_get_user_id(request),
            **validated,
        )
        if not updated:
            return Response({"detail": "Failed to update booking."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(BookingSerializer(updated).data)


class BookingAcceptView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        booking = accept_booking(booking_id, _get_user_id(request))
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        self._reconcile_b2b_booking(booking_id)
        return Response(BookingSerializer(booking).data)

    @staticmethod
    def _reconcile_b2b_booking(pms_booking_id: int) -> None:
        from apps.b2b.repository import get_b2b_booking_request_by_pms_booking_id
        from apps.b2b.hotel_booking_service import reconcile_booking_request

        booking_request = get_b2b_booking_request_by_pms_booking_id(pms_booking_id)
        if not booking_request:
            return
        try:
            reconcile_booking_request(booking_request)
        except Exception:
            logger.exception("Failed to reconcile B2B booking after PMS accept (pms_booking_id=%s)", pms_booking_id)


class BookingCancelView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        try:
            booking = cancel_booking(booking_id, _get_user_id(request))
        except IntegrityError as e:
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        BookingAcceptView._reconcile_b2b_booking(booking_id)
        return Response(BookingSerializer(booking).data)


class BookingCheckInView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        booking = check_in_booking(booking_id, _get_user_id(request))
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        BookingAcceptView._reconcile_b2b_booking(booking_id)
        return Response(BookingSerializer(booking).data)


class BookingCheckOutView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        booking = check_out_booking(booking_id, _get_user_id(request))
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        BookingAcceptView._reconcile_b2b_booking(booking_id)
        return Response(BookingSerializer(booking).data)


class BookingMoveView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        serializer = MoveBookingSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        booking = move_booking(
            booking_id,
            new_room_id=serializer.validated_data["new_room_id"],
            new_check_in=serializer.validated_data.get("new_check_in"),
            new_check_out=serializer.validated_data.get("new_check_out"),
            user_id=_get_user_id(request),
        )
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(BookingSerializer(booking).data)


class BookingMealPlanView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingSerializer()})
    def post(self, request, property_id, booking_id):
        serializer = MealPlanChangeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        booking = change_meal_plan(
            booking_id,
            serializer.validated_data["meal_plan"],
            _get_user_id(request),
        )
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(BookingSerializer(booking).data)


class BookingHistoryView(PMSBaseView):
    @swagger_auto_schema(responses={200: BookingHistorySerializer(many=True)})
    def get(self, request, property_id, booking_id):
        history = get_booking_history(booking_id)
        return Response(BookingHistorySerializer(history, many=True).data)


class BookingVoucherView(PMSBaseView):
    """GET  — read voucher_number for a booking.
    POST — regenerate / set voucher_number (only when booking is confirmed)."""

    @swagger_auto_schema(responses={200: BookingSerializer()})
    def get(self, request, property_id, booking_id):
        booking = get_booking(booking_id, property_id)
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(BookingSerializer(booking).data)

    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "voucher_number": openapi.Schema(type=openapi.TYPE_STRING, description="Optional custom voucher number. Auto-generated if omitted."),
            },
        ),
        responses={200: BookingSerializer()},
    )
    def post(self, request, property_id, booking_id):
        booking = get_booking(booking_id, property_id)
        if not booking:
            return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)

        if booking.get("status") != "confirmed":
            return Response(
                {"detail": "Voucher can only be set after booking is confirmed."},
                status=status.HTTP_409_CONFLICT,
            )

        voucher_number = request.data.get("voucher_number") or _generate_voucher_number(booking_id)
        updated = update_booking(booking_id, voucher_number=voucher_number)
        if not updated:
            return Response({"detail": "Failed to update voucher."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(BookingSerializer(updated).data)


class ReviewListCreateView(PMSBaseView):
    @swagger_auto_schema(responses={200: ReviewSerializer(many=True)})
    def get(self, request, property_id):
        rating = request.query_params.get("rating")
        is_complained = request.query_params.get("is_complained")
        reviews = list_reviews(
            property_id,
            rating=int(rating) if rating else None,
            is_complained=bool(is_complained) if is_complained else None,
        )
        return Response(ReviewSerializer(reviews, many=True).data)

    @swagger_auto_schema(responses={201: ReviewSerializer()})
    def post(self, request, property_id):
        serializer = ReviewSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            review = create_review(property_id=property_id, **serializer.validated_data)
        except IntegrityError as e:
            return Response({"detail": f"Database error: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if not review:
            return Response({"detail": "Failed to create review."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(ReviewSerializer(review).data, status=status.HTTP_201_CREATED)


class ReviewRespondView(PMSBaseView):
    @swagger_auto_schema(responses={200: ReviewSerializer()})
    def post(self, request, property_id, review_id):
        serializer = ReviewRespondSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        review = respond_to_review(review_id, serializer.validated_data["response"])
        if not review:
            return Response({"detail": "Review not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(ReviewSerializer(review).data)


class ReviewComplainView(PMSBaseView):
    @swagger_auto_schema(responses={200: ReviewSerializer()})
    def post(self, request, property_id, review_id):
        serializer = ReviewComplainSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        review = complain_review(review_id, serializer.validated_data["reason"])
        if not review:
            return Response({"detail": "Review not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(ReviewSerializer(review).data)


class AnalyticsView(PMSBaseView):
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("date_from", openapi.IN_QUERY, description="Start date", type=openapi.TYPE_STRING, format="date", required=True),
            openapi.Parameter("date_to", openapi.IN_QUERY, description="End date", type=openapi.TYPE_STRING, format="date", required=True),
            openapi.Parameter("metric", openapi.IN_QUERY, description="Chart metric", type=openapi.TYPE_STRING, enum=["check_ins", "revenue", "bookings", "occupancy"]),
            openapi.Parameter("category", openapi.IN_QUERY, description="Room category filter", type=openapi.TYPE_STRING),
            openapi.Parameter("floor", openapi.IN_QUERY, description="Floor filter", type=openapi.TYPE_STRING),
            openapi.Parameter("search", openapi.IN_QUERY, description="Room number search", type=openapi.TYPE_STRING),
        ],
        responses={200: AnalyticsResponseSerializer()},
    )
    def get(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)

        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = AnalyticsQuerySerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = get_analytics(
            property_id=property_id,
            date_from=serializer.validated_data["date_from"],
            date_to=serializer.validated_data["date_to"],
            metric=serializer.validated_data.get("metric", "revenue"),
            category=serializer.validated_data.get("category"),
            floor=serializer.validated_data.get("floor"),
            search=serializer.validated_data.get("search"),
        )

        return Response(AnalyticsResponseSerializer(data).data)


class AnalyticsExportView(PMSBaseView):
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("date_from", openapi.IN_QUERY, description="Start date", type=openapi.TYPE_STRING, format="date", required=True),
            openapi.Parameter("date_to", openapi.IN_QUERY, description="End date", type=openapi.TYPE_STRING, format="date", required=True),
            openapi.Parameter("metric", openapi.IN_QUERY, description="Chart metric", type=openapi.TYPE_STRING, enum=["check_ins", "revenue", "bookings", "occupancy"]),
            openapi.Parameter("category", openapi.IN_QUERY, description="Room category filter", type=openapi.TYPE_STRING),
            openapi.Parameter("floor", openapi.IN_QUERY, description="Floor filter", type=openapi.TYPE_STRING),
            openapi.Parameter("search", openapi.IN_QUERY, description="Room number search", type=openapi.TYPE_STRING),
        ],
        responses={200: openapi.Response(description="XLSX file", schema=openapi.Schema(type=openapi.TYPE_FILE))},
    )
    def get(self, request, property_id):
        org_id = _require_org(request)
        if not org_id:
            return Response({"detail": "Organization context required."}, status=status.HTTP_400_BAD_REQUEST)

        prop = get_property(property_id, organization_id=int(org_id))
        if not prop:
            return Response({"detail": "Property not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = AnalyticsQuerySerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = get_analytics(
            property_id=property_id,
            date_from=serializer.validated_data["date_from"],
            date_to=serializer.validated_data["date_to"],
            metric=serializer.validated_data.get("metric", "revenue"),
            category=serializer.validated_data.get("category"),
            floor=serializer.validated_data.get("floor"),
            search=serializer.validated_data.get("search"),
        )

        serialized = AnalyticsResponseSerializer(data).data
        rooms = serialized.get("rooms", [])
        date_from = serializer.validated_data["date_from"].isoformat()
        date_to = serializer.validated_data["date_to"].isoformat()

        wb = Workbook()
        ws = wb.active
        ws.title = "Room Analytics"

        currency = ""
        if rooms:
            currency = rooms[0].get("revenue", {}).get("currency", "")

        headers = ["Room", "Category", "Occupancy %", f"Revenue ({currency})", "Change %", f"ADR ({currency})", f"RevPAR ({currency})"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, room in enumerate(rooms, 2):
            ws.cell(row=row_idx, column=1, value=room.get("room_number"))
            ws.cell(row=row_idx, column=2, value=room.get("category"))
            ws.cell(row=row_idx, column=3, value=room.get("occupancy", {}).get("value"))
            ws.cell(row=row_idx, column=4, value=room.get("revenue", {}).get("value"))
            ws.cell(row=row_idx, column=5, value=room.get("revenue", {}).get("change_percent"))
            ws.cell(row=row_idx, column=6, value=room.get("adr", {}).get("value"))
            ws.cell(row=row_idx, column=7, value=room.get("revpar", {}).get("value"))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="analytics_{date_from}_{date_to}.xlsx"'
        return response
